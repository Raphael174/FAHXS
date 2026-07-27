""" 
@ author : Raphaël Aubry
"""

""" 
Script to estimate combustion chemistry using cantera
"""
#%%
import cantera as ct 
from pathlib import Path
import numpy as np 

#from input_data import hotgasProp
#ct.suppress_thermo_warnings()

#%%

base_path = Path(__file__).parent


def choose_fuel(fuel):

    if fuel == "POSF10325" :
        chem_mech_path = base_path / "A2highT.yaml"
        Hv_fuel = 360e3 #J/kg  https://web.stanford.edu/group/haiwanglab/HyChem/approach/Report_Jet_Fuel_Thermochemical_Properties_v6.pdf
        Y_fuel = {'POSF10325' : 1.}

    elif fuel == "H2" :
        chem_mech_path = base_path / "H2-O2_Burke2012.yaml"
        Hv_fuel = 0
        Y_fuel = {'H2' : 1.}

    elif fuel == "gasoline-E5" :
        chem_mech_path = base_path / "llnl_gasoline_323.yaml"
        Y_fuel = {'C2H5OH': np.float64(0.022521123344674744), 'C6H12-1': np.float64(0.05760037978051016),
                'C5H10-1': np.float64(0.020571564207325047),'IC4H8': np.float64(0.010971500910573358),
                'IC8H18': np.float64(0.6121661440853741), 'NC7H16': np.float64(0.2761692876715427)}
        Hv_fuel = 321.219e3 # J/kg

    elif fuel == "gasoline-E10" :
        chem_mech_path = base_path / "llnl_gasoline_323.yaml"
        Y_fuel = {'C2H5OH': np.float64(0.04648289377205732), 'C6H12-1': np.float64(0.059442690615179886),
                'C5H10-1': np.float64(0.021229532362564234), 'IC4H8': np.float64(0.01132241726003426),
                'IC8H18': np.float64(0.5936889008300161), 'NC7H16': np.float64(0.2678335651601482)}
        Hv_fuel = 313.664e3 # J/kg

    elif fuel == "diesel-C16H34" :
        # based on 10.1016/j.fuel.2017.07.009
        # DC surrogate, 
        chem_mech_path = base_path / "RenKokjohn_surrogate.yaml"
        Y_fuel = {'c16h34': np.float64(1)}
        Hv_fuel = 350e3 # J/kg, 10.4271/2008-01-1379
        

    else:
        raise("Error : fuel species not supported")
    
    return chem_mech_path, Y_fuel, Hv_fuel

#%%

class combustion_gas_solve :

    def __init__ (self, 
                  fuel, oxidizer,
                  OF,
                  p0,
                  T_g_init,
                  T_inj_LOX,
                  chem_mech_path, Y_fuel, Hv_fuel):
        
        self.fuel = fuel 
        self.oxidizer = oxidizer
        self.OF = OF 
        self.p0 = p0 
        self.T_g_init = T_g_init 
        self.T_inj_LOX = T_inj_LOX 

        self.phase = ct.Solution(chem_mech_path)
        self.Hv_fuel = Hv_fuel
        self.Y_fuel = Y_fuel

    def solve (self) :

        O = self.OF/(1 + self.OF)
        F = 1 - O

        """ENERGY LOSS LOX-->GOX"""

        # LOX TO GOX
        O2 = ct.Oxygen() # import oxygen 
        O2.TP = self.T_inj_LOX, self.p0 # set O2 to injection thermo
        self.h_ox_inj = O2.enthalpy_mass # fetch O2 onjection enthalpy
        O2.TP = self.T_g_init, self.p0 # set O2 to reaction state, assumed at T_pyro
        self.h_ox_pyro = O2.enthalpy_mass
        self.dH_ox = self.h_ox_pyro-self.h_ox_inj

        self.dH_tot = O*self.dH_ox + F*self.Hv_fuel


        """CHEMICAL EQUILIBRIUM CALCULATION"""

        if self.fuel == "gasoline-E5" or self.fuel == "gasoline-E10" or self.fuel == "diesel-C16H34":
            #! building the O2/gasoline mixture
            s = sum(self.Y_fuel.values())
            if s <= 0: raise ValueError("Y_fuel is empty.")
            Yf = {k:v/s for k,v in self.Y_fuel.items()}
            Y_mix = {k: (1.0/(1.0+self.OF))*Yf[k] for k in Yf}
            Y_mix[self.oxidizer] = self.OF/(1.0+self.OF)

        else:
            #! building the pure fuel/O2 mixture
            Y_mix = "O2" + ':' + str(self.OF) + ',' +  self.fuel + ': 1' 

        self.phase.TPY = self.T_g_init, self.p0, Y_mix

        self.phase.equilibrate('HP') # finding chem equilibrium

        self.phase.HP = self.phase.enthalpy_mass-self.dH_tot, self.p0 # remove energy loss
        self.phase.equilibrate('HP') # re-equilibriate mixture

    equilibrium_dh_gas_ON : bool = True
    def remove_energy (self, dh, updated_pressure, equilibrium_dh_gas_ON): 

        # enthalpy to remove
        self.dh = dh 
        self.p0=updated_pressure

        # equilibriate with removed energy
        self.phase.HP = self.phase.enthalpy_mass-self.dh, self.p0 # remove energy loss
        if equilibrium_dh_gas_ON==True:
            self.phase.equilibrate('HP') # re-equilibriate mixture

#%%

if __name__ == "__main__":
    """
    Finite-rate kinetics cooling study — hot combustion gas through a heat exchanger
    """
    import pandas as pd
    import matplotlib.pyplot as plt

    #%% ── Parameters ──────────────────────────────────────────────────────────────

    FUEL          = "diesel-C16H34"
    OXIDIZER      = "O2"
    OF            = 2.9
    P_COMB        = 5e5          # Pa
    T_IG          = 800 + 273.15 # K
    T_INJ_LOX     = 300          # K

    REMOVED_POWER = 441504       # W
    MASS_FLOW_RANGE = np.linspace(50, 150, 100, endpoint=True) * 1e-3   # kg/s

    L_HX  = 0.6   # m   — heat exchanger length (your only geometry input)
    MACH  = 0.3    # [-] — assumed Mach number at HX inlet
    N_STEPS = 500  # integration steps per case

    #%% ── Helper functions ────────────────────────────────────────────────────────

    def residence_time_from_mach(gas: ct.Solution, L_hx: float, mach: float):
        """Speed of sound → velocity → τ = L / v"""
        gamma      = gas.cp / gas.cv
        r_specific = ct.gas_constant / gas.mean_molecular_weight   # J/kg/K
        a          = np.sqrt(gamma * r_specific * gas.T)           # m/s
        v_gas      = mach * a
        return L_hx / v_gas, v_gas, a


    def finite_rate_hx(gas_inlet: ct.Solution,
                    mdot: float,
                    removed_power: float,
                    tau: float,
                    n_steps: int = 500) -> ct.SolutionArray:
        """
        Lagrangian parcel: IdealGasConstPressureReactor + prescribed wall heat flux.
        Returns SolutionArray over the HX integration (n_steps + 1 states).
        """
        q_dot_per_kg = removed_power / (mdot * tau)   # W/kg

        r    = ct.IdealGasConstPressureReactor(gas_inlet)
        sink = ct.Reservoir(ct.Solution(gas_inlet.source))
        wall = ct.Wall(r, sink, A=r.mass)
        wall.heat_flux = q_dot_per_kg      # W/m²  (area = r.mass → W/kg numerically)

        net = ct.ReactorNet([r])
        net.rtol = 1e-9
        net.atol = 1e-15

        states = ct.SolutionArray(gas_inlet, extra=['t'])
        states.append(r.thermo.state, t=0.0)

        dt = tau / n_steps
        for i in range(1, n_steps + 1):
            try:
                net.advance(i * dt)
                states.append(r.thermo.state, t=i * dt)
            except ct.CanteraError as e:
                print(f"  Integration stopped at step {i}/{n_steps}: {e}")
                break

        return states

    #%% ── Combustion setup (same for all mass flow points) ───────────────────────

    chem_mech_path, Y_fuel, Hv_fuel = choose_fuel(FUEL)

    # Solve combustion once to get τ (T_inlet is the same across all mdot)
    _ref_comb = combustion_gas_solve(
        fuel=FUEL, oxidizer=OXIDIZER, OF=OF, p0=P_COMB,
        T_g_init=T_IG, T_inj_LOX=T_INJ_LOX,
        chem_mech_path=chem_mech_path, Hv_fuel=Hv_fuel, Y_fuel=Y_fuel)
    _ref_comb.solve()

    tau, v_gas, a_sound = residence_time_from_mach(_ref_comb.phase, L_HX, MACH)

    print(f"HX inlet  :  T = {_ref_comb.phase.T:.0f} K")
    print(f"Sound speed: a = {a_sound:.1f} m/s")
    print(f"Flow speed : v = {v_gas:.1f} m/s  (M = {MACH})")
    print(f"Residence  : τ = {tau*1e3:.2f} ms\n")

    #%% ── Finite-rate sweep ────────────────────────────────────────────────────────

    dH_range = REMOVED_POWER / MASS_FLOW_RANGE   # J/kg

    results = {
        'mass_flow_kg_s' : MASS_FLOW_RANGE,
        'mass_flow_g_s'  : MASS_FLOW_RANGE * 1e3,
        'dH_removed_J_kg': dH_range,
        'T_g_K'          : np.zeros(len(MASS_FLOW_RANGE)),
        'cp_J_kgK'       : np.zeros(len(MASS_FLOW_RANGE)),
        'cv_J_kgK'       : np.zeros(len(MASS_FLOW_RANGE)),
        'lambda_W_mK'    : np.zeros(len(MASS_FLOW_RANGE)),
        'density_kg_m3'  : np.zeros(len(MASS_FLOW_RANGE)),
        'viscosity_Pa_s' : np.zeros(len(MASS_FLOW_RANGE)),
        'MW_kg_kmol'     : np.zeros(len(MASS_FLOW_RANGE)),
        'Pr'             : np.zeros(len(MASS_FLOW_RANGE)),
    }

    for i, mdot in enumerate(MASS_FLOW_RANGE):

        # fresh combustion solve every iteration (reactor will mutate the phase)
        comb = combustion_gas_solve(
            fuel=FUEL, oxidizer=OXIDIZER, OF=OF, p0=P_COMB,
            T_g_init=T_IG, T_inj_LOX=T_INJ_LOX,
            chem_mech_path=chem_mech_path, Hv_fuel=Hv_fuel, Y_fuel=Y_fuel)
        comb.solve()

        states = finite_rate_hx(
            gas_inlet     = comb.phase,
            mdot          = mdot,
            removed_power = REMOVED_POWER,
            tau           = tau,
            n_steps       = N_STEPS,
        )

        g = states[-1]   # HX exit state
        print(f"  mdot={mdot*1e3:6.1f} g/s | dH={dH_range[i]/1e3:7.1f} kJ/kg | T_exit={g.T:.0f} K")

        results['T_g_K'][i]          = g.T
        results['cp_J_kgK'][i]       = g.cp
        results['cv_J_kgK'][i]       = g.cv
        results['lambda_W_mK'][i]    = g.thermal_conductivity
        results['density_kg_m3'][i]  = g.density
        results['viscosity_Pa_s'][i] = g.viscosity
        results['MW_kg_kmol'][i]     = g.mean_molecular_weight
        results['Pr'][i]             = g.cp * g.viscosity / g.thermal_conductivity

    #%% ── Plots ───────────────────────────────────────────────────────────────────

    x   = results['mass_flow_g_s']
    x2  = results['dH_removed_J_kg'] / 1e3   # kJ/kg for secondary axis label

    plot_vars = [
        ('T_g_K',          r"$T_g$ [K]",            "Exit Temperature"),
        ('cp_J_kgK',       r"$c_p$ [J/kg·K]",       "Heat Capacity"),
        ('cv_J_kgK',       r"$c_v$ [J/kg·K]",       "Isochoric Heat Capacity"),
        ('lambda_W_mK',    r"$\lambda$ [W/m·K]",     "Thermal Conductivity"),
        ('density_kg_m3',  r"$\rho$ [kg/m³]",        "Density"),
        ('viscosity_Pa_s', r"$\mu$ [Pa·s]",          "Dynamic Viscosity"),
        ('MW_kg_kmol',     r"$MW$ [kg/kmol]",        "Mean Molecular Weight"),
        ('Pr',             r"Pr [-]",                "Prandtl Number"),
    ]

    fig, axes = plt.subplots(4, 2, figsize=(13, 16))
    fig.suptitle(
        f"Finite-rate kinetics — HX exit\n"
        f"Fuel: {FUEL}, O/F={OF}, M={MACH}, L={L_HX} m, τ={tau*1e3:.2f} ms",
        fontsize=12
    )

    for ax, (key, ylabel, title) in zip(axes.flat, plot_vars):
        ax.plot(x, results[key], color='tab:blue', linewidth=1.8)
        ax.set_xlabel(r"$\dot{m}$ [g/s]")
        ax.set_ylabel(ylabel)
        ax.set_title(title)
        ax.set_xlim(x[0], x[-1])
        ax.grid(which='major', color='#DDDDDD', linewidth=0.8)
        ax.grid(which='minor', color='#EEEEEE', linestyle=':', linewidth=0.5)
        ax.minorticks_on()

        # secondary x-axis showing specific enthalpy removal
        ax2 = ax.twiny()
        ax2.set_xlim(x2[0], x2[-1])
        ax2.set_xlabel(r"$\Delta h$ [kJ/kg]", fontsize=8, color='grey')
        ax2.tick_params(axis='x', labelsize=7, colors='grey')

    plt.tight_layout()
    plt.savefig(base_path / "finite_rate_hx_results.png", dpi=150)
    plt.show()

    #%% ── Save to Excel ───────────────────────────────────────────────────────────

    df = pd.DataFrame(results)

    col_labels = {
        'mass_flow_kg_s' : 'Mass flow [kg/s]',
        'mass_flow_g_s'  : 'Mass flow [g/s]',
        'dH_removed_J_kg': 'ΔH removed [J/kg]',
        'T_g_K'          : 'T_g [K]',
        'cp_J_kgK'       : 'cp [J/kg·K]',
        'cv_J_kgK'       : 'cv [J/kg·K]',
        'lambda_W_mK'    : 'λ [W/m·K]',
        'density_kg_m3'  : 'ρ [kg/m³]',
        'viscosity_Pa_s' : 'μ [Pa·s]',
        'MW_kg_kmol'     : 'MW [kg/kmol]',
        'Pr'             : 'Pr [-]',
    }
    df.rename(columns=col_labels, inplace=True)

    out_path = base_path / "finite_rate_hx_results.xlsx"

    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='finite_rate', index=False)

        ws = writer.sheets['finite_rate']
        ws.freeze_panes = 'A2'

        # header style
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill('solid', start_color='1F4E79')
        header_font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        data_font   = Font(name='Arial', size=10)

        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            ws.row_dimensions[1].height = 30

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font      = data_font
                cell.alignment = Alignment(horizontal='right')
                # scientific notation for very small numbers
                if cell.column_letter in ('I', 'J'):   # μ and λ columns
                    cell.number_format = '0.00E+00'
                else:
                    cell.number_format = '0.000'

        # metadata block
        meta_sheet = writer.book.create_sheet('parameters')
        meta = [
            ('Parameter', 'Value', 'Unit'),
            ('Fuel',       FUEL,            '—'),
            ('O/F',        OF,              '—'),
            ('p_comb',     P_COMB / 1e5,   'bar'),
            ('T_ignition', T_IG,            'K'),
            ('T_inj_LOX',  T_INJ_LOX,      'K'),
            ('Removed power', REMOVED_POWER,'W'),
            ('L_hx',       L_HX,           'm'),
            ('Mach',       MACH,           '—'),
            ('tau_ms',     round(tau*1e3, 3),'ms'),
            ('v_gas',      round(v_gas, 2), 'm/s'),
            ('a_sound',    round(a_sound, 2),'m/s'),
            ('n_steps',    N_STEPS,         '—'),
        ]
        for row in meta:
            meta_sheet.append(row)
        for cell in meta_sheet[1]:
            cell.font = Font(bold=True, name='Arial')
        for col in meta_sheet.columns:
            meta_sheet.column_dimensions[col[0].column_letter].width = 22

        # auto column width on data sheet
        for col in ws.columns:
            w = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = max(w + 2, 14)

    print(f"\nResults saved → {out_path}")